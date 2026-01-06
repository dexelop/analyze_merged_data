"""
회계 데이터 분석 스크립트 (10가지 분석)
- 기본 피벗, 거래처별, 증빙유형별, 월별 추이, 카드 현황
- 요일별 패턴, 금액구간별, 계정월별상세, 이상거래 탐지
"""
import json
import numpy as np
import pandas as pd
from pathlib import Path
from datetime import datetime

# 경로 설정
BASE_DIR = Path(__file__).parent.parent
INPUT_FILE = BASE_DIR / "input_merged_datas" / "더제이의원" / "result_2024_v01_20260106_225407.json"
OUTPUT_DIR = BASE_DIR / "output" / "더제이의원"
OUTPUT_DIR.mkdir(parents=True, exist_ok=True)

# 회사명 (파일명에서 추출하거나 지정)
COMPANY_NAME = "더제이의원"

# ============================================================
# 1. 데이터 로드
# ============================================================
print("1. 데이터 로드 중...")

# Excel 또는 JSON 파일 로드
if INPUT_FILE.suffix == '.xlsx':
    df = pd.read_excel(INPUT_FILE)
else:
    with open(INPUT_FILE, 'r', encoding='utf-8') as f:
        raw = json.load(f)
    df = pd.DataFrame(raw['data'])
df_original = df.copy()  # 원본 보존
print(f"   총 {len(df)}건 로드 완료")

# ============================================================
# 2. 파생 컬럼 생성
# ============================================================
print("2. 파생 컬럼 생성 중...")

# 2.1 소스유형: 전표번호 기준으로 vat/일반 구분
# - 카드미반영: 데이터소스가 '카드미반영'
# - 분개장(vat): 전표번호 >= 50000 (매입매출장 파생)
# - 분개장(일반): 전표번호 < 50000
def get_source_type(row):
    if row['데이터소스'] == '카드미반영':
        return '카드미반영'
    # 전표번호를 숫자로 변환 (빈값이나 문자열 처리)
    try:
        slip_no = int(row['전표번호']) if row['전표번호'] else 0
    except (ValueError, TypeError):
        slip_no = 0
    if slip_no >= 50000:
        return '분개장(vat)'
    return '분개장(일반)'

df['소스유형'] = df.apply(get_source_type, axis=1)

# 2.3 요일 (회계일자에서 추출)
df['회계일자_dt'] = pd.to_datetime(df['회계일자'], format='%Y%m%d', errors='coerce')
df['요일'] = df['회계일자_dt'].dt.dayofweek  # 0=월, 6=일
df['요일명'] = df['요일'].map({0:'월', 1:'화', 2:'수', 3:'목', 4:'금', 5:'토', 6:'일'})

# 2.4 금액구간
def get_amount_range(amt):
    abs_amt = abs(amt) if pd.notna(amt) else 0
    if abs_amt < 100_000:
        return "1_10만미만"
    if abs_amt < 500_000:
        return "2_10~50만"
    if abs_amt < 1_000_000:
        return "3_50~100만"
    if abs_amt < 5_000_000:
        return "4_100~500만"
    return "5_500만이상"

df['금액구간'] = df['순액'].apply(get_amount_range)

# 2.5 증빙유형명
evidence_names = {
    0: '수기',
    1: '현금조정',
    5: '결산분개',
    40: '원천세',
    86: '세금계산서',
    87: '영세율',
    88: '카드',
    88.5: '카드미반영',
    89: '현금영수증',
    90: '통장자동'
}
df['증빙유형명'] = df['증빙유형'].map(evidence_names).fillna(df['증빙유형'].astype(str))

print(f"   파생 컬럼 생성 완료")

# ============================================================
# 3. 기본 피벗 분석
# ============================================================
print("3. 기본 피벗 분석 중...")

# 계정코드별 고유 매핑 생성 (정렬용)
account_code_map = df.groupby('계정과목')['계정코드'].first().to_dict()

pivot_basic = df.pivot_table(
    index=['정렬순서', '손익분류', '계정과목'],
    columns='소스유형',
    values='순액',
    aggfunc='sum',
    fill_value=0
)

# 열 순서 정리 (분개장(vat), 분개장(일반), 분개장요약, 카드미반영, 총합계)
base_cols = ['분개장(vat)', '분개장(일반)']
pivot_basic = pivot_basic.reindex(columns=[c for c in base_cols if c in pivot_basic.columns], fill_value=0)

# 분개장 요약 컬럼 추가
if '분개장(vat)' in pivot_basic.columns and '분개장(일반)' in pivot_basic.columns:
    pivot_basic['분개장요약'] = pivot_basic['분개장(vat)'] + pivot_basic['분개장(일반)']
elif '분개장(vat)' in pivot_basic.columns:
    pivot_basic['분개장요약'] = pivot_basic['분개장(vat)']
elif '분개장(일반)' in pivot_basic.columns:
    pivot_basic['분개장요약'] = pivot_basic['분개장(일반)']

# 카드미반영 컬럼 추가
card_missing_sum = df[df['데이터소스'] == '카드미반영'].groupby(['정렬순서', '손익분류', '계정과목'])['순액'].sum()
pivot_basic['카드미반영'] = card_missing_sum.reindex(pivot_basic.index, fill_value=0)

# 총합계
pivot_basic['총합계'] = pivot_basic[['분개장요약', '카드미반영']].sum(axis=1)

# reset_index하여 정렬 (정렬순서 → 손익분류 → 계정코드 순)
pivot_basic = pivot_basic.reset_index()
pivot_basic['계정코드'] = pivot_basic['계정과목'].map(account_code_map)
pivot_basic = pivot_basic.sort_values(['정렬순서', '손익분류', '계정코드'])

# 컬럼 순서 정리 (계정코드는 제외)
col_order = ['정렬순서', '손익분류', '계정과목', '분개장(vat)', '분개장(일반)', '분개장요약', '카드미반영', '총합계']
pivot_basic = pivot_basic[[c for c in col_order if c in pivot_basic.columns]]

print(f"   기본 피벗: {len(pivot_basic)}행")

# ============================================================
# 3-2. 기본_거래처추가_피벗 (NEW)
# ============================================================
print("3-2. 기본_거래처추가_피벗 분석 중...")

# 거래처명이 비어있으면 "(미지정)"으로 처리
df['거래처명_filled'] = df['거래처명'].fillna('(미지정)').replace('', '(미지정)')

pivot_trader = df.pivot_table(
    index=['정렬순서', '손익분류', '계정과목', '거래처명_filled'],
    columns='소스유형',
    values='순액',
    aggfunc='sum',
    fill_value=0
)

# 열 순서 정리 (분개장(vat), 분개장(일반))
base_cols = ['분개장(vat)', '분개장(일반)']
pivot_trader = pivot_trader.reindex(columns=[c for c in base_cols if c in pivot_trader.columns], fill_value=0)

# 분개장 요약 컬럼 추가
if '분개장(vat)' in pivot_trader.columns and '분개장(일반)' in pivot_trader.columns:
    pivot_trader['분개장요약'] = pivot_trader['분개장(vat)'] + pivot_trader['분개장(일반)']
elif '분개장(vat)' in pivot_trader.columns:
    pivot_trader['분개장요약'] = pivot_trader['분개장(vat)']
elif '분개장(일반)' in pivot_trader.columns:
    pivot_trader['분개장요약'] = pivot_trader['분개장(일반)']
else:
    pivot_trader['분개장요약'] = 0

# 카드미반영 컬럼 추가
card_missing_trader = df[df['데이터소스'] == '카드미반영'].groupby(
    ['정렬순서', '손익분류', '계정과목', '거래처명_filled']
)['순액'].sum()
pivot_trader['카드미반영'] = card_missing_trader.reindex(pivot_trader.index, fill_value=0)

# 총합계
pivot_trader['총합계'] = pivot_trader[['분개장요약', '카드미반영']].sum(axis=1)

# 정렬: 정렬순서 → 손익분류 → 계정코드 → 분개장요약(내림차순)
pivot_trader = pivot_trader.reset_index()
pivot_trader['계정코드'] = pivot_trader['계정과목'].map(account_code_map)
pivot_trader = pivot_trader.sort_values(
    by=['정렬순서', '손익분류', '계정코드', '분개장요약'],
    ascending=[True, True, True, False]
)

# 컬럼명 변경 및 순서 정리 (계정코드 제외, MultiIndex 미사용)
pivot_trader = pivot_trader.rename(columns={'거래처명_filled': '거래처'})
col_order = ['정렬순서', '손익분류', '계정과목', '거래처', '분개장(vat)', '분개장(일반)', '분개장요약', '카드미반영', '총합계']
pivot_trader = pivot_trader[[c for c in col_order if c in pivot_trader.columns]]

print(f"   기본_거래처추가_피벗: {len(pivot_trader)}행")

# ============================================================
# 3-3. 기본_거래처_증빙유형 (NEW)
# ============================================================
print("3-3. 기본_거래처_증빙유형 분석 중...")

# 증빙유형 코드 순서 (0→1→5→40→86→87→88→88.5→89→90)
ev_type_order = [0, 1, 5, 40, 86, 87, 88, 88.5, 89, 90]

# 기본 피벗 생성
pivot_trader_ev = df.pivot_table(
    index=['정렬순서', '손익분류', '계정과목', '거래처명_filled', '증빙유형'],
    columns='소스유형',
    values='순액',
    aggfunc='sum',
    fill_value=0
)

# 열 순서 정리 (분개장(vat), 분개장(일반))
base_cols = ['분개장(vat)', '분개장(일반)']
pivot_trader_ev = pivot_trader_ev.reindex(columns=[c for c in base_cols if c in pivot_trader_ev.columns], fill_value=0)

# 분개장 요약 컬럼 추가
if '분개장(vat)' in pivot_trader_ev.columns and '분개장(일반)' in pivot_trader_ev.columns:
    pivot_trader_ev['분개장요약'] = pivot_trader_ev['분개장(vat)'] + pivot_trader_ev['분개장(일반)']
elif '분개장(vat)' in pivot_trader_ev.columns:
    pivot_trader_ev['분개장요약'] = pivot_trader_ev['분개장(vat)']
elif '분개장(일반)' in pivot_trader_ev.columns:
    pivot_trader_ev['분개장요약'] = pivot_trader_ev['분개장(일반)']
else:
    pivot_trader_ev['분개장요약'] = 0

# 카드미반영 컬럼 추가
card_missing_trader_ev = df[df['데이터소스'] == '카드미반영'].groupby(
    ['정렬순서', '손익분류', '계정과목', '거래처명_filled', '증빙유형']
)['순액'].sum()
pivot_trader_ev['카드미반영'] = card_missing_trader_ev.reindex(pivot_trader_ev.index, fill_value=0)

# 총합계
pivot_trader_ev['총합계'] = pivot_trader_ev[['분개장요약', '카드미반영']].sum(axis=1)

# reset_index for processing
pivot_trader_ev = pivot_trader_ev.reset_index()

# 거래처별 분개장요약 합계 계산 (정렬용)
trader_summary = pivot_trader_ev.groupby(
    ['정렬순서', '손익분류', '계정과목', '거래처명_filled']
)['분개장요약'].sum().reset_index()
trader_summary = trader_summary.rename(columns={'분개장요약': '거래처_분개장요약_합계'})

# 거래처 합계 조인
pivot_trader_ev = pivot_trader_ev.merge(trader_summary, on=['정렬순서', '손익분류', '계정과목', '거래처명_filled'])

# 모든 (계정과목, 거래처) 조합에 대해 모든 증빙유형 행 생성
unique_combinations = pivot_trader_ev[['정렬순서', '손익분류', '계정과목', '거래처명_filled', '거래처_분개장요약_합계']].drop_duplicates()

# 모든 조합 × 모든 증빙유형
from itertools import product
all_ev_types = pd.DataFrame({'증빙유형': ev_type_order})
full_index = pd.merge(unique_combinations.assign(key=1), all_ev_types.assign(key=1), on='key').drop('key', axis=1)

# 기존 데이터와 병합 (없는 조합은 0으로 채움)
pivot_trader_ev = full_index.merge(
    pivot_trader_ev,
    on=['정렬순서', '손익분류', '계정과목', '거래처명_filled', '증빙유형', '거래처_분개장요약_합계'],
    how='left'
).fillna(0)

# 증빙유형 순서 매핑
ev_type_sort_order = {v: i for i, v in enumerate(ev_type_order)}
pivot_trader_ev['증빙유형_순서'] = pivot_trader_ev['증빙유형'].map(ev_type_sort_order).fillna(999)

# 계정코드 추가 (정렬용)
pivot_trader_ev['계정코드'] = pivot_trader_ev['계정과목'].map(account_code_map)

# 정렬: 정렬순서 → 손익분류 → 계정코드 → 거래처(합계 내림차순) → 증빙유형(코드순)
pivot_trader_ev = pivot_trader_ev.sort_values(
    by=['정렬순서', '손익분류', '계정코드', '거래처_분개장요약_합계', '거래처명_filled', '증빙유형_순서'],
    ascending=[True, True, True, False, True, True]
)

# 불필요한 컬럼 제거
pivot_trader_ev = pivot_trader_ev.drop(columns=['거래처_분개장요약_합계', '증빙유형_순서', '계정코드'])

# 증빙유형 dict 변환
pivot_trader_ev['증빙유형'] = pivot_trader_ev['증빙유형'].map(evidence_names).fillna(pivot_trader_ev['증빙유형'].astype(str))

# 인덱스 재설정 (컬럼명 변경)
pivot_trader_ev = pivot_trader_ev.rename(columns={'거래처명_filled': '거래처'})

# 컬럼 순서 정리
col_order = ['정렬순서', '손익분류', '계정과목', '거래처', '증빙유형', '분개장(vat)', '분개장(일반)', '분개장요약', '카드미반영', '총합계']
pivot_trader_ev = pivot_trader_ev[[c for c in col_order if c in pivot_trader_ev.columns]]

# 인덱스 설정하지 않음 (MultiIndex merged cell 문제 방지)
# pivot_trader_ev = pivot_trader_ev.set_index(['정렬순서', '손익분류', '계정과목', '거래처', '증빙유형'])

print(f"   기본_거래처_증빙유형: {len(pivot_trader_ev)}행")

# ============================================================
# 3-4. total_월별추이_가로 (Option B: 소스유형 × 월별 컬럼)
# ============================================================
print("3-4. total_월별추이_가로 분석 중...")

# 월별 + 소스유형별 피벗
monthly_wide = df.pivot_table(
    index=['정렬순서', '손익분류', '계정과목', '거래처명_filled', '증빙유형'],
    columns=['월', '소스유형'],
    values='순액',
    aggfunc='sum',
    fill_value=0
)

# 컬럼 평탄화 (01_분개장(vat), 01_분개장(일반), ...)
monthly_wide.columns = [f'{month}_{source}' for month, source in monthly_wide.columns]
monthly_wide = monthly_wide.reset_index()

# 계정코드 추가하여 정렬
monthly_wide['계정코드'] = monthly_wide['계정과목'].map(account_code_map)

# 거래처별 합계 계산 (정렬용)
value_cols = [c for c in monthly_wide.columns if c not in ['정렬순서', '손익분류', '계정과목', '거래처명_filled', '증빙유형', '계정코드']]
monthly_wide['거래처_합계'] = monthly_wide[value_cols].sum(axis=1)

# 증빙유형 순서 매핑
monthly_wide['증빙유형_순서'] = monthly_wide['증빙유형'].map(ev_type_sort_order).fillna(999)

# 정렬: 정렬순서 → 손익분류 → 계정코드 → 거래처합계(내림차순) → 증빙유형순서
monthly_wide = monthly_wide.sort_values(
    by=['정렬순서', '손익분류', '계정코드', '거래처_합계', '증빙유형_순서'],
    ascending=[True, True, True, False, True]
)

# 정리용 컬럼 제거
monthly_wide = monthly_wide.drop(columns=['계정코드', '거래처_합계', '증빙유형_순서'])

# 컬럼명 변경
monthly_wide = monthly_wide.rename(columns={'거래처명_filled': '거래처'})

# 증빙유형 dict 변환
monthly_wide['증빙유형'] = monthly_wide['증빙유형'].map(evidence_names).fillna(monthly_wide['증빙유형'].astype(str))

# 월별 컬럼 순서 정렬 (소스유형별 그룹 → 월 순서)
# 결과: 01_vat, 02_vat, ..., 12_vat, 01_일반, 02_일반, ..., 12_일반, 01_카드미반영, ...
base_cols = ['정렬순서', '손익분류', '계정과목', '거래처', '증빙유형']
def get_col_sort_key(col_name):
    month = col_name[:2]  # '01', '02', etc.
    if 'vat' in col_name:
        source_order = 0
    elif '일반' in col_name:
        source_order = 1
    else:  # 카드미반영
        source_order = 2
    return (source_order, month)

month_cols = sorted([c for c in monthly_wide.columns if c not in base_cols], key=get_col_sort_key)
monthly_wide = monthly_wide[base_cols + month_cols]

# 합계 컬럼 추가
monthly_wide['합계'] = monthly_wide[month_cols].sum(axis=1)

print(f"   total_월별추이_가로: {len(monthly_wide)}행, {len(monthly_wide.columns)}컬럼")

# ============================================================
# 3-5. total_월별추이_세로 (Option C: 월을 행으로)
# ============================================================
print("3-5. total_월별추이_세로 분석 중...")

# 월별 + 소스유형별 피벗 (월을 행에 포함)
monthly_long = df.pivot_table(
    index=['정렬순서', '손익분류', '계정과목', '거래처명_filled', '증빙유형', '월'],
    columns='소스유형',
    values='순액',
    aggfunc='sum',
    fill_value=0
).reset_index()

# 컬럼 순서 정리
base_cols_long = ['분개장(vat)', '분개장(일반)']
for col in base_cols_long:
    if col not in monthly_long.columns:
        monthly_long[col] = 0

# 분개장요약
monthly_long['분개장요약'] = monthly_long.get('분개장(vat)', 0) + monthly_long.get('분개장(일반)', 0)

# 카드미반영 컬럼
if '카드미반영' not in monthly_long.columns:
    monthly_long['카드미반영'] = 0

# 총합계
monthly_long['총합계'] = monthly_long['분개장요약'] + monthly_long['카드미반영']

# 계정코드 추가하여 정렬
monthly_long['계정코드'] = monthly_long['계정과목'].map(account_code_map)

# 거래처별 총합계 계산 (정렬용)
trader_total = monthly_long.groupby(['정렬순서', '손익분류', '계정과목', '거래처명_filled'])['총합계'].sum().reset_index()
trader_total = trader_total.rename(columns={'총합계': '거래처_총합계'})
monthly_long = monthly_long.merge(trader_total, on=['정렬순서', '손익분류', '계정과목', '거래처명_filled'])

# 증빙유형 순서 매핑
monthly_long['증빙유형_순서'] = monthly_long['증빙유형'].map(ev_type_sort_order).fillna(999)

# 정렬: 정렬순서 → 손익분류 → 계정코드 → 거래처총합계(내림차순) → 증빙유형순서 → 월
monthly_long = monthly_long.sort_values(
    by=['정렬순서', '손익분류', '계정코드', '거래처_총합계', '거래처명_filled', '증빙유형_순서', '월'],
    ascending=[True, True, True, False, True, True, True]
)

# 정리용 컬럼 제거
monthly_long = monthly_long.drop(columns=['계정코드', '거래처_총합계', '증빙유형_순서'])

# 컬럼명 변경
monthly_long = monthly_long.rename(columns={'거래처명_filled': '거래처'})

# 증빙유형 dict 변환
monthly_long['증빙유형'] = monthly_long['증빙유형'].map(evidence_names).fillna(monthly_long['증빙유형'].astype(str))

# 컬럼 순서 정리
col_order_long = ['정렬순서', '손익분류', '계정과목', '거래처', '증빙유형', '월', '분개장(vat)', '분개장(일반)', '분개장요약', '카드미반영', '총합계']
monthly_long = monthly_long[[c for c in col_order_long if c in monthly_long.columns]]

print(f"   total_월별추이_세로: {len(monthly_long)}행")

# ============================================================
# 3-6. total_월별추이_가로_빈도 (거래 횟수 기준)
# ============================================================
print("3-6. total_월별추이_가로_빈도 분석 중...")

# 월별 + 소스유형별 피벗 (거래 횟수)
monthly_wide_cnt = df.pivot_table(
    index=['정렬순서', '손익분류', '계정과목', '거래처명_filled', '증빙유형'],
    columns=['월', '소스유형'],
    values='순액',
    aggfunc='count',
    fill_value=0
)

# 컬럼 평탄화
monthly_wide_cnt.columns = [f'{month}_{source}' for month, source in monthly_wide_cnt.columns]
monthly_wide_cnt = monthly_wide_cnt.reset_index()

# 계정코드 추가하여 정렬
monthly_wide_cnt['계정코드'] = monthly_wide_cnt['계정과목'].map(account_code_map)

# 거래처별 합계 계산 (정렬용)
value_cols_cnt = [c for c in monthly_wide_cnt.columns if c not in ['정렬순서', '손익분류', '계정과목', '거래처명_filled', '증빙유형', '계정코드']]
monthly_wide_cnt['거래처_합계'] = monthly_wide_cnt[value_cols_cnt].sum(axis=1)

# 증빙유형 순서 매핑
monthly_wide_cnt['증빙유형_순서'] = monthly_wide_cnt['증빙유형'].map(ev_type_sort_order).fillna(999)

# 정렬
monthly_wide_cnt = monthly_wide_cnt.sort_values(
    by=['정렬순서', '손익분류', '계정코드', '거래처_합계', '증빙유형_순서'],
    ascending=[True, True, True, False, True]
)

# 정리용 컬럼 제거
monthly_wide_cnt = monthly_wide_cnt.drop(columns=['계정코드', '거래처_합계', '증빙유형_순서'])

# 컬럼명 변경
monthly_wide_cnt = monthly_wide_cnt.rename(columns={'거래처명_filled': '거래처'})

# 증빙유형 dict 변환
monthly_wide_cnt['증빙유형'] = monthly_wide_cnt['증빙유형'].map(evidence_names).fillna(monthly_wide_cnt['증빙유형'].astype(str))

# 월별 컬럼 순서 정렬 (소스유형별 그룹 → 월 순서)
base_cols_cnt = ['정렬순서', '손익분류', '계정과목', '거래처', '증빙유형']
month_cols_cnt = sorted([c for c in monthly_wide_cnt.columns if c not in base_cols_cnt], key=get_col_sort_key)
monthly_wide_cnt = monthly_wide_cnt[base_cols_cnt + month_cols_cnt]

# 합계 컬럼 추가
monthly_wide_cnt['합계'] = monthly_wide_cnt[month_cols_cnt].sum(axis=1)

print(f"   total_월별추이_가로_빈도: {len(monthly_wide_cnt)}행, {len(monthly_wide_cnt.columns)}컬럼")

# ============================================================
# 3-7. total_월별추이_세로_빈도 (거래 횟수 기준)
# ============================================================
print("3-7. total_월별추이_세로_빈도 분석 중...")

# 월별 + 소스유형별 피벗 (거래 횟수, 월을 행에 포함)
monthly_long_cnt = df.pivot_table(
    index=['정렬순서', '손익분류', '계정과목', '거래처명_filled', '증빙유형', '월'],
    columns='소스유형',
    values='순액',
    aggfunc='count',
    fill_value=0
).reset_index()

# 컬럼 순서 정리
for col in ['분개장(vat)', '분개장(일반)']:
    if col not in monthly_long_cnt.columns:
        monthly_long_cnt[col] = 0

# 분개장요약
monthly_long_cnt['분개장요약'] = monthly_long_cnt.get('분개장(vat)', 0) + monthly_long_cnt.get('분개장(일반)', 0)

# 카드미반영 컬럼
if '카드미반영' not in monthly_long_cnt.columns:
    monthly_long_cnt['카드미반영'] = 0

# 총합계
monthly_long_cnt['총합계'] = monthly_long_cnt['분개장요약'] + monthly_long_cnt['카드미반영']

# 계정코드 추가하여 정렬
monthly_long_cnt['계정코드'] = monthly_long_cnt['계정과목'].map(account_code_map)

# 거래처별 총합계 계산 (정렬용)
trader_total_cnt = monthly_long_cnt.groupby(['정렬순서', '손익분류', '계정과목', '거래처명_filled'])['총합계'].sum().reset_index()
trader_total_cnt = trader_total_cnt.rename(columns={'총합계': '거래처_총합계'})
monthly_long_cnt = monthly_long_cnt.merge(trader_total_cnt, on=['정렬순서', '손익분류', '계정과목', '거래처명_filled'])

# 증빙유형 순서 매핑
monthly_long_cnt['증빙유형_순서'] = monthly_long_cnt['증빙유형'].map(ev_type_sort_order).fillna(999)

# 정렬
monthly_long_cnt = monthly_long_cnt.sort_values(
    by=['정렬순서', '손익분류', '계정코드', '거래처_총합계', '거래처명_filled', '증빙유형_순서', '월'],
    ascending=[True, True, True, False, True, True, True]
)

# 정리용 컬럼 제거
monthly_long_cnt = monthly_long_cnt.drop(columns=['계정코드', '거래처_총합계', '증빙유형_순서'])

# 컬럼명 변경
monthly_long_cnt = monthly_long_cnt.rename(columns={'거래처명_filled': '거래처'})

# 증빙유형 dict 변환
monthly_long_cnt['증빙유형'] = monthly_long_cnt['증빙유형'].map(evidence_names).fillna(monthly_long_cnt['증빙유형'].astype(str))

# 컬럼 순서 정리
col_order_long_cnt = ['정렬순서', '손익분류', '계정과목', '거래처', '증빙유형', '월', '분개장(vat)', '분개장(일반)', '분개장요약', '카드미반영', '총합계']
monthly_long_cnt = monthly_long_cnt[[c for c in col_order_long_cnt if c in monthly_long_cnt.columns]]

print(f"   total_월별추이_세로_빈도: {len(monthly_long_cnt)}행")

# ============================================================
# 4. 거래처별 분석 (판관비 중심)
# ============================================================
print("4. 거래처별 분석 중...")

df_pangwan = df[df['손익분류'] == '판관비'].copy()

# 계정과목별 거래처 TOP 10
trade_top = df_pangwan.groupby(['계정과목', '거래처명'])['순액'].sum().reset_index()
trade_top = trade_top.sort_values(['계정과목', '순액'], ascending=[True, False])
trade_top['rank'] = trade_top.groupby('계정과목')['순액'].rank(method='first', ascending=False)
trade_top10 = trade_top[trade_top['rank'] <= 10].drop(columns=['rank'])

print(f"   거래처 분석: {len(trade_top10)}건 (TOP 10 per 계정)")

# ============================================================
# 5. 증빙유형별 분석
# ============================================================
print("5. 증빙유형별 분석 중...")

evidence_analysis = df.pivot_table(
    index=['손익분류', '계정과목'],
    columns='증빙유형명',
    values='순액',
    aggfunc='sum',
    fill_value=0
)

evidence_analysis['합계'] = evidence_analysis.sum(axis=1)

# 증빙률 계산 (세금계산서+카드+현금영수증 / 전체)
vat_cols = ['세금계산서', '카드', '현금영수증']
evidence_analysis['증빙금액'] = evidence_analysis[[c for c in vat_cols if c in evidence_analysis.columns]].sum(axis=1)
evidence_analysis['증빙률'] = (evidence_analysis['증빙금액'] / evidence_analysis['합계'].replace(0, 1) * 100).round(1)

print(f"   증빙유형 분석: {len(evidence_analysis)}행")

# ============================================================
# 6. 월별 추이 분석
# ============================================================
print("6. 월별 추이 분석 중...")

monthly_trend = df.pivot_table(
    index='손익분류',
    columns='월',
    values='순액',
    aggfunc='sum',
    fill_value=0
)

monthly_trend['합계'] = monthly_trend.sum(axis=1)
monthly_trend['평균'] = monthly_trend.iloc[:, :-1].mean(axis=1).round(0)

# 정렬순서 기준으로 정렬
sort_order = df.groupby('손익분류')['정렬순서'].first().sort_values()
monthly_trend = monthly_trend.reindex(sort_order.index)

print(f"   월별 추이: {len(monthly_trend)}행")

# ============================================================
# 7. 카드 현황 분석
# ============================================================
print("7. 카드 현황 분석 중...")

df_card = df[df['증빙유형'].isin([88, 88.5])].copy()

if len(df_card) > 0:
    df_card['전표상태'] = df_card['전표상태'].fillna('없음')
    df_card['공제구분'] = df_card['공제구분'].fillna('없음')

    card_status = df_card.pivot_table(
        index='계정과목',
        columns=['공제구분', '전표상태'],
        values='순액',
        aggfunc='sum',
        fill_value=0
    )
    print(f"   카드 현황: {len(card_status)}행")
else:
    card_status = pd.DataFrame()
    print("   카드 데이터 없음")

# ============================================================
# 8. 카드미반영 상세
# ============================================================
print("8. 카드미반영 상세 분석 중...")

df_card_missing = df[df['증빙유형'] == 88.5].copy()

if len(df_card_missing) > 0:
    # 컬럼 목록 (업태, 업종을 계정과목 다음에 배치)
    detail_cols = ['회계일자', '거래처명', '순액', '공제구분', '전표상태', '계정과목']
    # 업태, 업종 컬럼이 있으면 추가
    if '업태' in df_card_missing.columns:
        detail_cols.insert(detail_cols.index('계정과목') + 1, '업태')
    if '업종' in df_card_missing.columns:
        detail_cols.insert(detail_cols.index('업태') + 1 if '업태' in detail_cols else detail_cols.index('계정과목') + 1, '업종')

    # 존재하는 컬럼만 선택
    available_cols = [c for c in detail_cols if c in df_card_missing.columns]
    card_missing_detail = df_card_missing[available_cols].copy()

    # 업태, 업종 컬럼이 없으면 빈 컬럼 추가
    if '업태' not in card_missing_detail.columns:
        card_missing_detail.insert(card_missing_detail.columns.get_loc('계정과목') + 1, '업태', '')
    if '업종' not in card_missing_detail.columns:
        card_missing_detail.insert(card_missing_detail.columns.get_loc('업태') + 1, '업종', '')

    card_missing_detail = card_missing_detail.sort_values('회계일자')
    print(f"   카드미반영: {len(card_missing_detail)}건, 총 {card_missing_detail['순액'].sum():,.0f}원")
else:
    card_missing_detail = pd.DataFrame()
    print("   카드미반영 없음")

# ============================================================
# 9. 요일별 패턴 분석 (NEW)
# ============================================================
print("9. 요일별 패턴 분석 중...")

# 요일별 × 손익분류
weekday_pattern = df.pivot_table(
    index='요일명',
    columns='손익분류',
    values='순액',
    aggfunc=['sum', 'count'],
    fill_value=0
)

# 요일 순서대로 정렬
weekday_order = ['월', '화', '수', '목', '금', '토', '일']
weekday_pattern = weekday_pattern.reindex([d for d in weekday_order if d in weekday_pattern.index])

# 요일별 총계
weekday_summary = df.groupby('요일명').agg({
    '순액': ['sum', 'count', 'mean']
}).reset_index()
weekday_summary.columns = ['요일', '총금액', '건수', '평균금액']
weekday_summary['요일순서'] = weekday_summary['요일'].map({d: i for i, d in enumerate(weekday_order)})
weekday_summary = weekday_summary.sort_values('요일순서').drop(columns=['요일순서'])

print(f"   요일별 패턴: {len(weekday_summary)}행")

# ============================================================
# 10. 금액구간별 분석 (NEW)
# ============================================================
print("10. 금액구간별 분석 중...")

# 금액구간 × 손익분류
amount_range_analysis = df.pivot_table(
    index='금액구간',
    columns='손익분류',
    values='순액',
    aggfunc=['sum', 'count'],
    fill_value=0
)

# 금액구간별 요약
amount_summary = df.groupby('금액구간').agg({
    '순액': ['sum', 'count', 'mean']
}).reset_index()
amount_summary.columns = ['금액구간', '총금액', '건수', '평균금액']
amount_summary = amount_summary.sort_values('금액구간')

print(f"   금액구간별: {len(amount_summary)}행")

# ============================================================
# 11. 계정과목별 월별 상세 (NEW)
# ============================================================
print("11. 계정과목별 월별 상세 분석 중...")

account_monthly = df.pivot_table(
    index=['정렬순서', '손익분류', '계정과목'],
    columns='월',
    values='순액',
    aggfunc='sum',
    fill_value=0
)

account_monthly['합계'] = account_monthly.sum(axis=1)
account_monthly = account_monthly.sort_index(level=0)

print(f"   계정월별: {len(account_monthly)}행")

# ============================================================
# 12. 이상 거래 탐지 (NEW)
# ============================================================
print("12. 이상 거래 탐지 중...")

anomalies = []

# 12.1 금액 이상 탐지 (계정과목별 Z-score)
account_stats = df.groupby('계정과목')['순액'].agg(['mean', 'std'])

for account in account_stats.index:
    mean_val = account_stats.loc[account, 'mean']
    std_val = account_stats.loc[account, 'std']

    if std_val > 0:
        account_data = df[df['계정과목'] == account]
        for _, row in account_data.iterrows():
            z_score = (row['순액'] - mean_val) / std_val
            if abs(z_score) > 3:  # 3 표준편차 초과
                anomalies.append({
                    '유형': '금액이상',
                    '계정과목': account,
                    '거래처명': row.get('거래처명', ''),
                    '회계일자': row.get('회계일자', ''),
                    '금액': row['순액'],
                    '평균': round(mean_val, 0),
                    'Z-score': round(z_score, 2),
                    '비고': f'평균 대비 {abs(z_score):.1f}σ 이탈'
                })

# 12.2 마이너스 금액 탐지 (비용에서 음수)
expense_categories = ['판관비', '매출원가', '영업외비용']
negative_expenses = df[(df['손익분류'].isin(expense_categories)) & (df['순액'] < 0)]

for _, row in negative_expenses.iterrows():
    anomalies.append({
        '유형': '마이너스',
        '계정과목': row['계정과목'],
        '거래처명': row.get('거래처명', ''),
        '회계일자': row.get('회계일자', ''),
        '금액': row['순액'],
        '평균': 0,
        'Z-score': 0,
        '비고': '비용 계정에서 음수 (환불?)'
    })

# 12.3 월별 급변 탐지
monthly_by_account = df.groupby(['계정과목', '월'])['순액'].sum().reset_index()
monthly_by_account = monthly_by_account.sort_values(['계정과목', '월'])

for account in monthly_by_account['계정과목'].unique():
    acc_data = monthly_by_account[monthly_by_account['계정과목'] == account].copy()
    if len(acc_data) < 2:
        continue

    acc_data['prev'] = acc_data['순액'].shift(1)
    acc_data['change_rate'] = (acc_data['순액'] - acc_data['prev']) / acc_data['prev'].replace(0, 1)

    for _, row in acc_data.iterrows():
        if pd.isna(row['change_rate']):
            continue
        if row['change_rate'] > 2:  # 200% 이상 급증
            anomalies.append({
                '유형': '급증',
                '계정과목': account,
                '거래처명': '',
                '회계일자': f"2024{row['월']}",
                '금액': row['순액'],
                '평균': row['prev'],
                'Z-score': 0,
                '비고': f"전월 대비 {row['change_rate']*100:.0f}% 증가"
            })
        elif row['change_rate'] < -0.5:  # 50% 이상 급감
            anomalies.append({
                '유형': '급감',
                '계정과목': account,
                '거래처명': '',
                '회계일자': f"2024{row['월']}",
                '금액': row['순액'],
                '평균': row['prev'],
                'Z-score': 0,
                '비고': f"전월 대비 {row['change_rate']*100:.0f}% 감소"
            })

anomaly_df = pd.DataFrame(anomalies) if anomalies else pd.DataFrame()
print(f"   이상 거래: {len(anomaly_df)}건 탐지")

# ============================================================
# 13. Excel 출력
# ============================================================
print("13. Excel 출력 중...")

from openpyxl.utils import get_column_letter
from openpyxl.styles import numbers

# 파일명에 타임스탬프 추가 (mm-dd-hh-mm)
timestamp = datetime.now().strftime("%m-%d-%H-%M")
excel_path = OUTPUT_DIR / f"분석결과_{timestamp}.xlsx"

with pd.ExcelWriter(excel_path, engine='openpyxl') as writer:
    # 원본 데이터 (맨 앞)
    df_original.to_excel(writer, sheet_name='원본데이터', index=False)

    # 1. 기본 피벗 - index=False로 병합 셀 문제 방지
    pivot_basic.to_excel(writer, sheet_name='기본피벗', index=False)

    # 1-2. 기본_거래처추가_피벗 (NEW) - index=False로 병합 셀 문제 방지
    pivot_trader.to_excel(writer, sheet_name='기본_거래처추가_피벗', index=False)

    # 1-3. 기본_거래처_증빙유형 (NEW) - index=False로 병합 셀 문제 방지
    pivot_trader_ev.to_excel(writer, sheet_name='기본_거래처_증빙유형', index=False)

    # 1-4. total_월별추이_가로 (소스유형 × 월별 컬럼)
    monthly_wide.to_excel(writer, sheet_name='total_월별추이_가로', index=False)

    # 1-5. total_월별추이_세로 (월을 행으로)
    monthly_long.to_excel(writer, sheet_name='total_월별추이_세로', index=False)

    # 1-6. total_월별추이_가로_빈도 (거래 횟수)
    monthly_wide_cnt.to_excel(writer, sheet_name='total_월별추이_가로_빈도', index=False)

    # 1-7. total_월별추이_세로_빈도 (거래 횟수)
    monthly_long_cnt.to_excel(writer, sheet_name='total_월별추이_세로_빈도', index=False)

    # 2. 월별 추이
    monthly_trend.to_excel(writer, sheet_name='월별추이')

    # 3. 계정월별상세
    account_monthly.to_excel(writer, sheet_name='계정월별')

    # 4. 거래처 TOP 10
    trade_top10.to_excel(writer, sheet_name='거래처TOP', index=False)

    # 5. 증빙유형별
    evidence_analysis.to_excel(writer, sheet_name='증빙유형별')

    # 6. 카드 현황
    if len(card_status) > 0:
        card_status.to_excel(writer, sheet_name='카드현황')

    # 7. 카드미반영 상세
    if len(card_missing_detail) > 0:
        card_missing_detail.to_excel(writer, sheet_name='카드미반영', index=False)

    # 8. 요일별 패턴
    weekday_summary.to_excel(writer, sheet_name='요일별', index=False)

    # 9. 금액구간별
    amount_summary.to_excel(writer, sheet_name='금액구간별', index=False)

    # 10. 이상 거래
    if len(anomaly_df) > 0:
        anomaly_df.to_excel(writer, sheet_name='이상거래', index=False)

    # 서식 적용: 열너비 자동 조정 + 금액 형식
    workbook = writer.book

    # 금액 관련 컬럼명 (이 컬럼들에 #,##0 형식 적용)
    money_keywords = ['금액', '순액', '합계', '평균', '총금액', '평균금액', '분개장', '카드미반영', '총합계',
                      '증빙금액', '세금계산서', '카드', '현금영수증', '수기', '결산분개', '원천세',
                      '영세율', '통장자동', '현금조정']

    for sheet_name in workbook.sheetnames:
        ws = workbook[sheet_name]

        # 열너비 자동 조정
        for col_idx, column_cells in enumerate(ws.columns, 1):
            max_length = 0
            column_letter = get_column_letter(col_idx)

            for cell in column_cells:
                try:
                    cell_value = str(cell.value) if cell.value is not None else ""
                    cell_length = len(cell_value.encode('utf-8'))  # 한글 고려
                    # 한글은 2바이트로 계산
                    adjusted_length = sum(2 if ord(c) > 127 else 1 for c in cell_value)
                    if adjusted_length > max_length:
                        max_length = adjusted_length
                except:
                    pass

            # 최소 8, 최대 50
            adjusted_width = min(max(max_length + 2, 8), 50)
            ws.column_dimensions[column_letter].width = adjusted_width

        # 금액 형식 적용 (#,##0)
        # 첫 번째 행(헤더) 확인
        header_row = list(ws.iter_rows(min_row=1, max_row=1, values_only=True))[0]

        for col_idx, header in enumerate(header_row, 1):
            if header is None:
                continue

            # 금액 관련 컬럼인지 확인
            is_money_col = any(kw in str(header) for kw in money_keywords)

            if is_money_col:
                column_letter = get_column_letter(col_idx)
                for row_idx in range(2, ws.max_row + 1):
                    cell = ws[f"{column_letter}{row_idx}"]
                    if isinstance(cell.value, (int, float)):
                        cell.number_format = '#,##0'

print(f"   Excel 저장: {excel_path}")

# ============================================================
# 14. JSON 출력
# ============================================================
print("14. JSON 출력 중...")

def df_to_dict(df):
    """DataFrame을 JSON 직렬화 가능한 dict로 변환"""
    result = df.copy()
    if isinstance(result.index, pd.MultiIndex):
        result = result.reset_index()
    for col in result.columns:
        if result[col].dtype in ['int64', 'float64']:
            result[col] = result[col].apply(
                lambda x: int(x) if pd.notna(x) and x == int(x)
                else float(x) if pd.notna(x) else None
            )
    return result.to_dict(orient='records')

json_output = {
    "meta": {
        "회사명": COMPANY_NAME,
        "기간": "2024",
        "총건수": len(df),
        "생성일시": datetime.now().isoformat()
    },
    "기본피벗": df_to_dict(pivot_basic.reset_index()),
    "기본_거래처추가_피벗": df_to_dict(pivot_trader.reset_index()),
    "기본_거래처_증빙유형": df_to_dict(pivot_trader_ev.reset_index()),
    "월별추이": df_to_dict(monthly_trend.reset_index()),
    "계정월별": df_to_dict(account_monthly.reset_index()),
    "거래처TOP": df_to_dict(trade_top10),
    "증빙유형별": df_to_dict(evidence_analysis.reset_index()),
    "요일별": df_to_dict(weekday_summary),
    "금액구간별": df_to_dict(amount_summary),
    "카드미반영": df_to_dict(card_missing_detail) if len(card_missing_detail) > 0 else [],
    "이상거래": df_to_dict(anomaly_df) if len(anomaly_df) > 0 else []
}

json_path = OUTPUT_DIR / f"분석결과_{timestamp}.json"
with open(json_path, 'w', encoding='utf-8') as f:
    json.dump(json_output, f, ensure_ascii=False, indent=2)

print(f"   JSON 저장: {json_path}")

# ============================================================
# 15. 요약 출력
# ============================================================
print("\n" + "="*60)
print("분석 완료! (10가지 분석)")
print("="*60)

print(f"\n[데이터 요약]")
print(f"  - 총 건수: {len(df):,}건")
print(f"  - 기간: 2024년 {df['월'].min()}월 ~ {df['월'].max()}월")

print(f"\n[손익 요약]")
for idx, row in monthly_trend.iterrows():
    print(f"  - {idx}: {row['합계']:,.0f}원")

print(f"\n[분석 결과]")
print(f"  - 기본 피벗: {len(pivot_basic)}행")
print(f"  - 기본_거래처추가_피벗: {len(pivot_trader)}행")
print(f"  - 기본_거래처_증빙유형: {len(pivot_trader_ev)}행")
print(f"  - 월별 추이: {len(monthly_trend)}행")
print(f"  - 계정월별: {len(account_monthly)}행")
print(f"  - 거래처 TOP: {len(trade_top10)}건")
print(f"  - 증빙유형별: {len(evidence_analysis)}행")
print(f"  - 카드미반영: {len(card_missing_detail)}건")
print(f"  - 요일별: {len(weekday_summary)}행")
print(f"  - 금액구간별: {len(amount_summary)}행")
print(f"  - 이상 거래: {len(anomaly_df)}건")

print(f"\n[출력 파일]")
print(f"  - Excel: {excel_path}")
print(f"  - JSON: {json_path}")

print(f"\n[Excel 시트 목록]")
sheets = [
    "원본데이터", "기본피벗", "기본_거래처추가_피벗", "기본_거래처_증빙유형",
    "월별추이", "계정월별", "거래처TOP", "증빙유형별", "카드현황", "카드미반영",
    "요일별", "금액구간별", "이상거래"
]
for i, sheet in enumerate(sheets, 1):
    print(f"  {i:2}. {sheet}")
